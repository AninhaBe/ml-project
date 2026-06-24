from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analyzer import AnaliseConcorrencia
from .margin import CalculoMargem
from .pdf_parser import ProdutoFornecedor

logger = logging.getLogger(__name__)


@dataclass
class LinhaResultado:
    produto_fornec_nome: str
    produto_fornec_codigo: str
    produto_fornec_preco: float | None
    catalog_id: str | None
    catalog_name: str
    status_match: str  # codigo_exato | fuzzy_nome | vision_llm | vision_review | nao_encontrado
    confianca_match: float
    n_concorrentes: int
    n_full: int
    preco_min: float
    preco_mediana: float
    preco_max: float
    lider_full_preco: float | None
    lider_full_seller: str
    visitas_30d: int
    margem_pct: float | None
    margem_pct_premium: float | None
    pv_alvo: float | None
    lucro_alvo: float | None
    score_oportunidade: float
    bandeiras: str
    veredicto: str  # APROVADO | AVALIAR | REJEITAR | DESCARTADO | REVIEW
    permalink_lider: str


VERDICT_COLORS = {
    "APROVADO": "C6EFCE",   # verde
    "AVALIAR": "FFEB9C",    # amarelo
    "REJEITAR": "FFC7CE",   # vermelho
    "REVIEW": "BDD7EE",     # azul
    "DESCARTADO": "D9D9D9", # cinza
}


def gerar_excel(linhas: list[LinhaResultado], out_path) -> object:
    """Gera Excel formatado. Retorna o path final do arquivo."""
    df = pd.DataFrame([_linha_to_dict(l) for l in linhas])
    df = df.sort_values(by="score_oportunidade", ascending=False, ignore_index=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not out_path.suffix:
        out_path = out_path.with_suffix(".xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Oportunidades", index=False)
        _formatar(writer.book["Oportunidades"])
    return out_path


def _linha_to_dict(l: LinhaResultado) -> dict:
    return {
        "Produto Fornecedor": l.produto_fornec_nome,
        "Código Fornec.": l.produto_fornec_codigo or "",
        "Custo (R$)": l.produto_fornec_preco,
        "Catálogo ML": l.catalog_id or "",
        "Nome no ML": l.catalog_name,
        "Status Match": l.status_match,
        "Confiança": round(l.confianca_match, 2),
        "# Concorrentes": l.n_concorrentes,
        "# Full": l.n_full,
        "Preço Min": l.preco_min,
        "Preço Mediana": l.preco_mediana,
        "Preço Max": l.preco_max,
        "Líder Full Preço": l.lider_full_preco,
        "Líder Full Seller": l.lider_full_seller,
        "Visitas 30d": l.visitas_30d,
        "Margem % (clássico)": _pct(l.margem_pct),
        "Margem % (premium)": _pct(l.margem_pct_premium),
        "PV Alvo": l.pv_alvo,
        "Lucro Alvo": l.lucro_alvo,
        "Score": round(l.score_oportunidade, 1),
        "Bandeiras": l.bandeiras,
        "Veredicto": l.veredicto,
        "Permalink": l.permalink_lider,
    }


def _pct(v: float | None) -> str:
    if v is None:
        return ""
    return f"{v * 100:.1f}%"


def _formatar(ws) -> None:
    # Header em negrito + fundo cinza
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-width simples
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Colorir linha por Veredicto
    headers = [c.value for c in ws[1]]
    try:
        idx_veredicto = headers.index("Veredicto") + 1
    except ValueError:
        return

    for row in ws.iter_rows(min_row=2):
        veredicto = row[idx_veredicto - 1].value
        cor = VERDICT_COLORS.get(str(veredicto))
        if cor:
            fill = PatternFill("solid", fgColor=cor)
            for cell in row:
                cell.fill = fill


def montar_linha(
    produto: ProdutoFornecedor,
    analise: AnaliseConcorrencia | None,
    margem_classico: CalculoMargem | None,
    margem_premium: CalculoMargem | None,
    margem_minima: float,
    score: float,
    status_match: str,
    confianca_match: float,
    veredicto_motivo: str = "",
) -> LinhaResultado:
    lider_full = analise.lider_full if analise else None
    lider = lider_full or (analise.lider_geral if analise else None)
    permalink = ""
    if analise and analise.catalog_product_id:
        permalink = f"https://www.mercadolivre.com.br/p/{analise.catalog_product_id}"
    elif lider:
        permalink = f"https://www.mercadolivre.com.br/p/{lider.item_id}"
    # Veredicto
    if status_match == "nao_encontrado":
        veredicto = "DESCARTADO"
    elif status_match == "vision_review":
        veredicto = "REVIEW"
    elif analise and analise.catalogo_fantasma:
        veredicto = "DESCARTADO"
    elif status_match == "sem_criterio":
        veredicto = "AVALIAR"  # encontrou mas demanda baixa/seller único
    elif margem_classico is None or analise is None:
        veredicto = "REVIEW"
    elif margem_classico.margem_pct < margem_minima:
        veredicto = "REJEITAR"
    elif analise and "marca_vende_direto" in analise.bandeiras:
        veredicto = "AVALIAR"
    elif analise and "demanda_baixa" in analise.bandeiras:
        veredicto = "AVALIAR"
    else:
        veredicto = "APROVADO"

    return LinhaResultado(
        produto_fornec_nome=produto.nome,
        produto_fornec_codigo=produto.codigo or "",
        produto_fornec_preco=produto.preco,
        catalog_id=analise.catalog_product_id if analise else None,
        catalog_name=analise.catalog_name if analise else "",
        status_match=status_match,
        confianca_match=confianca_match,
        n_concorrentes=analise.n_concorrentes if analise else 0,
        n_full=analise.n_full if analise else 0,
        preco_min=analise.preco_min if analise else 0.0,
        preco_mediana=analise.preco_mediana if analise else 0.0,
        preco_max=analise.preco_max if analise else 0.0,
        lider_full_preco=lider_full.preco if lider_full else None,
        lider_full_seller=lider_full.seller_nickname if lider_full else "",
        visitas_30d=analise.visitas_total_30d if analise else 0,
        margem_pct=margem_classico.margem_pct if margem_classico else None,
        margem_pct_premium=margem_premium.margem_pct if margem_premium else None,
        pv_alvo=margem_classico.preco_venda if margem_classico else None,
        lucro_alvo=margem_classico.lucro if margem_classico else None,
        score_oportunidade=score,
        bandeiras=", ".join(analise.bandeiras) if analise else "",
        veredicto=veredicto,
        permalink_lider=permalink,
    )


def nome_excel_default():
    from .config import OUTPUTS_DIR
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUTS_DIR / f"analise_{ts}.xlsx"
